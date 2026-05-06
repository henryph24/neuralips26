"""Build an xlsx file recording the per-bibitem academic-source verification status.

Source of truth: main.tex thebibliography (lines 662-949). Verification queries
were issued to Exa academic search; results captured in the rows below.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROWS = [
    # (idx, bibkey, first_author, year_in_cite, claimed_venue, search_status, found_url, source_kind, notes)
    (1,  "vaswani2017attention",   "Vaswani",    2017, "NeurIPS 2017",        "FOUND", "https://arxiv.org/abs/1706.03762", "academic_conf", "Verified at proceedings.neurips.cc/paper_files/paper/2017"),
    (2,  "bommasani2021foundation","Bommasani",  2021, "arXiv:2108.07258",    "FOUND", "https://arxiv.org/abs/2108.07258", "academic_preprint", "Stanford CRFM report; arXiv preprint"),
    (3,  "devlin2019bert",         "Devlin",     2019, "NAACL 2019",          "FOUND", "https://aclanthology.org/N19-1423/", "academic_conf", "ACL Anthology, NAACL-HLT 2019, pp.4171-4186"),
    (4,  "brown2020gpt3",          "Brown",      2020, "NeurIPS 2020",        "FOUND", "https://papers.neurips.cc/paper/2020/hash/1457c0d6bfcb4967418bfb8ac142f64a-Abstract.html", "academic_conf", "NeurIPS 2020 official proceedings"),
    (5,  "dosovitskiy2021vit",     "Dosovitskiy",2021, "ICLR 2021",           "FOUND", "https://iclr.cc/virtual/2021/poster/3013", "academic_conf", "ICLR 2021 poster; arXiv 2010.11929"),
    (6,  "goswami2024moment",      "Goswami",    2024, "ICML 2024 (PMLR 235)","FOUND", "https://proceedings.mlr.press/v235/goswami24a.html", "academic_conf", "ICML 2024, pp.16115-16152"),
    (7,  "das2024timesfm",         "Das",        2024, "ICML 2024",           "FOUND", "https://proceedings.mlr.press/v235/das24c.html", "academic_conf", "ICML 2024 (PMLR v235), pp.10148-10167"),
    (8,  "ansari2024chronos",      "Ansari",     2024, "arXiv:2403.07815",    "FOUND", "https://arxiv.org/abs/2403.07815", "academic_preprint", "arXiv preprint; published at TMLR 2024"),
    (9,  "liu2024timer",           "Liu",        2024, "ICML 2024",           "FOUND", "https://proceedings.mlr.press/v235/liu24cb.html", "academic_conf", "ICML 2024 (PMLR v235), pp.32369-32399"),
    (10, "liu2025timerxl",         "Liu",        2025, "ICLR 2025",           "FOUND", "https://iclr.cc/virtual/2025/poster/30062", "academic_conf", "ICLR 2025 poster"),
    (11, "woo2024moirai",          "Woo",        2024, "ICML 2024",           "FOUND", "https://icml.cc/virtual/2024/oral/35515", "academic_conf", "ICML 2024 oral"),
    (12, "liu2025moiraimoe",       "Liu",        2025, "ICML 2025",           "FOUND", "https://proceedings.mlr.press/v267/liu25an.html", "academic_conf", "ICML 2025 (PMLR v267), pp.38940-38962"),
    (13, "lee2024units",           "Lee",        2024, "NeurIPS 2024",        "AUTHOR_MISMATCH", "https://openreview.net/forum?id=nBOdYBptWW", "academic_conf", "BIB ERROR: real authors are Gao, Koker, Queen, Hartvigsen, Tsiligkaridis, Zitnik (Harvard/MIT-LL); NOT Lee, Liu, Sahoo. Paper IS real (NeurIPS 2024) but bib metadata is wrong."),
    (14, "rasul2024lagllama",      "Rasul",      2024, "arXiv:2310.08278",    "FOUND", "https://arxiv.org/abs/2310.08278", "academic_preprint", "arXiv preprint, multiple revisions"),
    (15, "shi2024timemoe",         "Shi",        2024, "arXiv:2409.16040",    "FOUND", "https://arxiv.org/abs/2409.16040", "academic_preprint", "arXiv preprint; also accepted ICLR 2025 Spotlight"),
    (16, "ekambaram2024ttm",       "Ekambaram",  2024, "arXiv:2401.03955",    "FOUND", "https://arxiv.org/abs/2401.03955", "academic_preprint", "arXiv preprint; also NeurIPS 2024"),
    (17, "nie2023patchtst",        "Nie",        2023, "ICLR 2023",           "FOUND", "https://iclr.cc/virtual/2023/poster/10876", "academic_conf", "ICLR 2023 poster; arXiv 2211.14730"),
    (18, "liu2024itransformer",    "Liu",        2024, "ICLR 2024",           "FOUND", "https://proceedings.iclr.cc/paper_files/paper/2024/hash/2ea18fdc667e0ef2ad82b2b4d65147ad-Abstract-Conference.html", "academic_conf", "ICLR 2024 official proceedings"),
    (19, "wu2023timesnet",         "Wu",         2023, "ICLR 2023",           "FOUND", "https://iclr.cc/virtual/2023/poster/11976", "academic_conf", "ICLR 2023 poster"),
    (20, "hu2022lora",             "Hu",         2022, "ICLR 2022",           "FOUND", "https://openreview.net/pdf?id=nZeVKeeFYf9", "academic_conf", "ICLR 2022 OpenReview"),
    (21, "houlsby2019adapters",    "Houlsby",    2019, "ICML 2019",           "FOUND", "https://proceedings.mlr.press/v97/houlsby19a.html", "academic_conf", "ICML 2019 (PMLR v97), pp.2790-2799"),
    (22, "zeng2023dlinear",        "Zeng",       2023, "AAAI 2023",           "FOUND", "https://ojs.aaai.org/index.php/AAAI/article/view/26317", "academic_conf", "AAAI-23 official, vol 37, pp.11121-11128"),
    (23, "wang2022adamix",         "Wang",       2022, "EMNLP 2022",          "FOUND", "https://aclanthology.org/2022.emnlp-main.388/", "academic_conf", "EMNLP 2022, pp.5744-5760"),
    (24, "fedus2022switch",        "Fedus",      2022, "JMLR 23(120)",        "FOUND", "https://jmlr.org/papers/v23/21-0998.html", "academic_journal", "JMLR 23(120):1-39, 2022"),
    (25, "kraskov2004ksg",         "Kraskov",    2004, "Phys. Rev. E 69(6)",  "FOUND", "https://journals.aps.org/pre/abstract/10.1103/PhysRevE.69.066138", "academic_journal", "Phys. Rev. E 69, 066138 (2004)"),
    (26, "kim2021revin",           "Kim",        2021, "ICLR 2021",           "VENUE_MISMATCH", "https://iclr.cc/virtual/2022/poster/6034", "academic_conf", "BIB CONCERN: paper appears to be ICLR 2022 (poster id 6034) not ICLR 2021. Paper is real but year/venue may be off by one."),
    (27, "zhang2025remoe",         "Zhang",      2025, "ICLR 2025",           "AUTHOR_MISMATCH", "https://proceedings.iclr.cc/paper_files/paper/2025/hash/94dc604e115237a7f4a758b3146cd976-Abstract-Conference.html", "academic_conf", "BIB ERROR: real authors are Wang, Zhu, Chen (Tsinghua); NOT Zhang, Yuan, Li, Stoica. Paper IS real (ICLR 2025) but bib metadata is wrong."),
    (28, "zhou2022expertchoice",   "Zhou",       2022, "NeurIPS 2022",        "FOUND", "https://proceedings.neurips.cc/paper_files/paper/2022/hash/2f00ecd787b432c1d36f3de9800728eb-Abstract-Conference.html", "academic_conf", "NeurIPS 2022 official proceedings"),
    (29, "qiao2025msft",           "Qiao",       2025, "NeurIPS 2025",        "FOUND", "https://neurips.cc/virtual/2025/poster/118294", "academic_conf", "NeurIPS 2025 poster (Dec 5, 2025)"),
    (30, "li2025trace",            "Li",         2025, "arXiv:2503.16991",    "FOUND", "https://arxiv.org/abs/2503.16991", "academic_preprint", "arXiv preprint, Tsinghua/HKU"),
    (31, "gupta2024beyondlora",    "Gupta",      2024, "arXiv:2409.11302",    "FOUND", "https://arxiv.org/abs/2409.11302", "academic_preprint", "arXiv preprint; also NeurIPS 2024 workshop on Time Series"),
    (32, "benechehab2025adapts",   "Benechehab", 2025, "arXiv:2502.10235",    "FOUND", "https://proceedings.mlr.press/v267/benechehab25a.html", "academic_preprint_and_conf", "arXiv 2502.10235; also published at ICML 2025 (PMLR v267)"),
    (33, "zhao2025prune",          "Zhao",       2025, "NeurIPS 2025",        "FOUND", "https://nips.cc/virtual/2025/loc/san-diego/poster/116420", "academic_conf", "NeurIPS 2025 poster; arXiv 2505.23195"),
    (34, "faw2025icf",             "Faw",        2025, "ICML 2025",           "FOUND", "https://proceedings.mlr.press/v267/faw25b.html", "academic_conf", "ICML 2025 (PMLR v267), pp.16355-16374"),
    (35, "zhang2025template",      "Zhang",      2025, "NeurIPS 2025",        "FOUND", "https://neurips.cc/virtual/2025/poster/117228", "academic_conf", "NeurIPS 2025 poster; method named TEMPLATE"),
    (36, "woo2024gifteval",        "Woo",        2024, "arXiv:2410.10393",    "AUTHOR_INCOMPLETE", "https://arxiv.org/abs/2410.10393", "academic_preprint", "BIB CONCERN: bib uses 'Woo et al.' shorthand, but actual first author is Taha Aksu (Woo is 2nd). Paper IS real (arXiv preprint)."),
    (37, "zoph2022stmoe",          "Zoph",       2022, "arXiv:2202.08906",    "FOUND", "https://arxiv.org/abs/2202.08906", "academic_preprint", "arXiv preprint, Google Brain"),
    (38, "fan2023dishts",          "Fan",        2023, "AAAI 2023",           "FOUND", "https://arxiv.org/abs/2302.14829", "academic_conf", "AAAI 2023; arXiv 2302.14829 (note paper title is 'Dish-TS' with hyphen)"),
    (39, "liu2022nonstationary",   "Liu",        2022, "NeurIPS 2022",        "FOUND", "https://proceedings.neurips.cc/paper_files/paper/2022/hash/4054556fcaa934b0bf76da52cf4f92cb-Abstract-Conference.html", "academic_conf", "NeurIPS 2022 official proceedings"),
    (40, "zhou2021informer",       "Zhou",       2021, "AAAI 2021",           "FOUND", "https://ojs.aaai.org/index.php/AAAI/article/view/17325", "academic_conf", "AAAI 2021, vol 35, pp.11106-11115"),
    (41, "godahewa2021monash",     "Godahewa",   2021, "NeurIPS D&B 2021",    "FOUND", "https://neurips.cc/virtual/2021/29865", "academic_conf", "NeurIPS Datasets & Benchmarks Track 2021; arXiv 2105.06643"),
    (42, "wen2023tssurvey",        "Wen",        2023, "IJCAI 2023",          "FOUND", "https://www.ijcai.org/proceedings/2023/759", "academic_conf", "IJCAI 2023 Survey Track, pp.6778-6786"),
    (43, "howard2017mobilenets",   "Howard",     2017, "arXiv:1704.04861",    "FOUND", "https://arxiv.org/abs/1704.04861", "academic_preprint", "arXiv preprint, Google"),
    (44, "he2016resnet",           "He",         2016, "CVPR 2016",           "FOUND", "https://openaccess.thecvf.com/content_cvpr_2016/html/He_Deep_Residual_Learning_CVPR_2016_paper.html", "academic_conf", "CVPR 2016 open access; arXiv 1512.03385"),
    (45, "hu2018senet",            "Hu",         2018, "CVPR 2018",           "FOUND", "https://openaccess.thecvf.com/content_cvpr_2018/html/Hu_Squeeze-and-Excitation_Networks_CVPR_2018_paper", "academic_conf", "CVPR 2018, pp.7132-7141"),
    (46, "lv2025aoe",              "Lv",         2025, "ICML 2025",           "FOUND", "https://proceedings.mlr.press/v267/lv25b.html", "academic_conf", "ICML 2025 (PMLR v267), pp.41667-41681; arXiv 2501.13074"),
    (47, "cheng2025ermoe",         "Cheng",      2025, "arXiv:2511.10971",    "FOUND", "https://arxiv.org/abs/2511.10971", "academic_preprint", "arXiv preprint, USC"),
    (48, "liu2026routingfree",     "Liu",        2026, "arXiv:2604.00801",    "FOUND", "https://arxiv.org/abs/2604.00801", "academic_preprint", "arXiv preprint; LMU Munich/UCLA, Apr 2026"),
    (49, "cover2006elements",      "Cover",      2006, "Wiley book (2nd ed)", "FOUND", "https://www.wiley.com/en-us/Elements%2Bof%2BInformation%2BTheory%252C%2B2nd%2BEdition-p-9780471241959", "academic_book", "Wiley-Interscience textbook, ISBN 978-0-471-24195-9"),
    (50, "tishby2000ib",           "Tishby",     2000, "Allerton 2000",       "FOUND", "https://arxiv.org/abs/physics/0004057", "academic_preprint_and_conf", "arXiv physics/0004057; preceded Allerton publication"),
    (51, "liu2026timers1",         "Liu",        2026, "arXiv:2603.04791",    "FOUND", "https://arxiv.org/abs/2603.04791", "academic_preprint", "arXiv preprint, Tsinghua/ByteDance, Mar 2026"),
    (52, "berthelier2026revin",    "Berthelier", 2026, "arXiv:2603.11869",    "FOUND", "https://arxiv.org/abs/2603.11869", "academic_preprint", "arXiv preprint, EDF/Inria, Mar 2026"),
    (53, "wang2026myth",           "Wang",       2026, "arXiv:2604.09780",    "FOUND", "https://arxiv.org/abs/2604.09780", "academic_preprint", "arXiv preprint, Johns Hopkins, Apr 2026"),
    (54, "zou2025ibnorm",          "Zou",        2025, "arXiv:2510.25262",    "FOUND", "https://arxiv.org/abs/2510.25262", "academic_preprint", "arXiv preprint; under ICLR 2026 review"),
    (55, "chi2022representation",  "Chi",        2022, "NeurIPS 2022",        "FOUND", "https://proceedings.neurips.cc/paper_files/paper/2022/hash/df4f371f1f89ec8ba5014b3310578048-Abstract-Conference.html", "academic_conf", "NeurIPS 2022 official proceedings"),
    (56, "hua2025inputaware",      "Hua",        2025, "ACM MM 2025",         "FOUND", "https://arxiv.org/abs/2510.16448", "academic_conf", "ACM MM '25 (Dublin, Oct 2025); arXiv 2510.16448"),
    (57, "reimers2019sentencebert","Reimers",    2019, "EMNLP 2019",          "FOUND", "https://aclanthology.org/D19-1410/", "academic_conf", "EMNLP-IJCNLP 2019, pp.3982-3992"),
]

EXPECTED_COUNT = 57

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bib Verification"

    headers = ["#", "BibKey", "First Author", "Year (cite)", "Claimed Venue", "Status",
               "Verified URL", "Source Kind", "Notes"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    bad_fill = PatternFill("solid", fgColor="FFD6D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")

    for r in ROWS:
        ws.append(list(r))
        row_idx = ws.max_row
        status = r[5]
        fill = ok_fill if status == "FOUND" else (bad_fill if "MISMATCH" in status else warn_fill)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    # Column widths
    widths = [4, 28, 14, 11, 24, 18, 70, 28, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    found    = sum(1 for r in ROWS if r[5] == "FOUND")
    aut_mm   = sum(1 for r in ROWS if r[5] == "AUTHOR_MISMATCH")
    aut_inc  = sum(1 for r in ROWS if r[5] == "AUTHOR_INCOMPLETE")
    venue_mm = sum(1 for r in ROWS if r[5] == "VENUE_MISMATCH")

    ws2["A1"] = "Bibliography verification summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Total rows in xlsx", len(ROWS)])
    ws2.append(["Expected count (from \\bibitem grep on main.tex)", EXPECTED_COUNT])
    ws2.append(["Counts match", "YES" if len(ROWS) == EXPECTED_COUNT else "NO"])
    ws2.append([])
    ws2.append(["FOUND (clean)", found])
    ws2.append(["AUTHOR_MISMATCH (paper exists, wrong authors in bib)", aut_mm])
    ws2.append(["AUTHOR_INCOMPLETE (paper exists, et al. shorthand misses true first author)", aut_inc])
    ws2.append(["VENUE_MISMATCH (paper exists, year/venue off-by-one)", venue_mm])
    ws2.append(["Sum of all categories", found + aut_mm + aut_inc + venue_mm])
    ws2.append([])
    ws2.append(["All cited papers exist as academic sources?",
                "YES (every entry resolves to a real conference/journal/arXiv preprint/textbook)"])
    ws2.append(["Source kinds present",
                "academic_conf, academic_journal, academic_preprint, academic_book, academic_preprint_and_conf"])

    ws2.column_dimensions["A"].width = 70
    ws2.column_dimensions["B"].width = 20

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "bib_verification.xlsx")
    wb.save(out)
    print(f"Wrote {out}")
    print(f"Rows in xlsx: {len(ROWS)}; expected from grep: {EXPECTED_COUNT}; match: {len(ROWS) == EXPECTED_COUNT}")
    print(f"FOUND={found}, AUTHOR_MISMATCH={aut_mm}, AUTHOR_INCOMPLETE={aut_inc}, VENUE_MISMATCH={venue_mm}")

if __name__ == "__main__":
    main()
